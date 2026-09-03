"""
metrics/parse_agent_results.py
Parse all agent JSON result files into a single formatted Excel workbook.
Run from project root: python metrics/parse_agent_results.py
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Style helpers ─────────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row: int, col: int, value: str, bg: str = "1F3864"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border_thin()
    return c

def data_cell(ws, row: int, col: int, value, bg: str = "FFFFFF",
              bold: bool = False, center: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=10)
    c.fill      = fill(bg)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True
    )
    c.border    = border_thin()
    return c

# ── Colors ────────────────────────────────────────────────────────────────────

BG_BUG      = "FFE0E0"   # red   — bug detected
BG_CLEAN    = "E8F5E9"   # green — no bug
BG_HEADER   = "FFF2CC"   # yellow — section header
BG_META     = "EEF2FF"   # light blue — run metadata
BG_ALT      = "F9F9F9"   # alternating row

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    results_dir = Path("results/agent")
    output_path = Path("results/agent_evaluations.xlsx")

    if not results_dir.exists():
        print("No results/agent directory found.")
        return

    files = sorted(results_dir.glob("*.json"))
    if not files:
        print("No agent JSON result files found.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: All evaluations combined ─────────────────────────────────────
    ws_all = wb.create_sheet("All Evaluations")
    ws_all.sheet_view.showGridLines = False
    ws_all.freeze_panes = "A3"

    # Title
    ws_all.merge_cells("A1:H1")
    t = ws_all["A1"]
    t.value     = "AI Agent — All Evaluation Results"
    t.font      = Font(name="Arial", bold=True, size=14, color="1F3864")
    t.fill      = fill(BG_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.border    = border_thin()
    ws_all.row_dimensions[1].height = 28

    headers = [
        "Run", "App", "Mode",
        "Test ID", "Endpoint",
        "Bug Detected", "Verdict", "LLM Reasoning"
    ]
    for i, h in enumerate(headers):
        header_cell(ws_all, 2, i+1, h)
    ws_all.row_dimensions[2].height = 20

    all_row = 3

    # ── One sheet per result file ──────────────────────────────────────────────
    for f in files:
        try:
            with open(f) as fp:
                data = json.load(fp)
        except Exception as e:
            print(f"Warning: could not load {f.name}: {e}")
            continue

        meta        = data.get("meta", {})
        target      = meta.get("target", "unknown")
        mode        = meta.get("mode", "black")
        evaluations = data.get("evaluations", [])
        summary     = data.get("summary", {})

        if not evaluations:
            print(f"Skipping {f.name} — no evaluations")
            continue

        # Create per-file sheet
        sheet_name = f"{target}_{mode}"[:31]   # Excel sheet name limit
        # Make unique if duplicate
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + f"_{len(existing)}"

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        # Sheet title
        ws.merge_cells("A1:G1")
        st = ws["A1"]
        st.value     = f"Agent Run: {target.upper()} — {mode.upper()} mode"
        st.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
        st.fill      = fill(BG_HEADER)
        st.alignment = Alignment(horizontal="center", vertical="center")
        st.border    = border_thin()
        ws.row_dimensions[1].height = 24

        # Metadata row
        bugs_found = summary.get("bugs_detected", 0)
        fp_count   = summary.get("false_positives", 0)
        tokens     = meta.get("token_usage", 0)
        meta_str   = (
            f"Tests executed: {summary.get('tests_executed',0)}  |  "
            f"Bugs detected: {bugs_found}  |  "
            f"False positives: {fp_count}  |  "
            f"Tokens: {tokens:,}  |  "
            f"Timestamp: {meta.get('timestamp','')[:19]}"
        )
        ws.merge_cells("A2:G2")
        m = ws["A2"]
        m.value     = meta_str
        m.font      = Font(name="Arial", size=9, italic=True)
        m.fill      = fill(BG_META)
        m.alignment = Alignment(horizontal="left", vertical="center")
        m.border    = border_thin()
        ws.row_dimensions[2].height = 16

        # Headers
        sheet_headers = [
            "Test ID", "Endpoint", "Bug Detected",
            "Status Received", "Passed", "Verdict", "LLM Reasoning"
        ]
        for i, h in enumerate(sheet_headers):
            header_cell(ws, 3, i+1, h)
        ws.row_dimensions[3].height = 20

        # Data rows
        for i, ev in enumerate(evaluations):
            row    = i + 4
            bug    = ev.get("bug_detected", False)
            passed = ev.get("passed", False)
            bg     = BG_BUG if bug else (BG_CLEAN if passed else BG_ALT)

            data_cell(ws, row, 1, ev.get("test_id", ""),       bg=BG_ALT, bold=True)
            data_cell(ws, row, 2, ev.get("endpoint", ""),      bg=bg)
            data_cell(ws, row, 3, "YES" if bug else "no",      bg=bg, bold=bug, center=True)
            data_cell(ws, row, 4, ev.get("status_received", ""), bg=bg, center=True)
            data_cell(ws, row, 5, "✓" if passed else "✗",      bg=bg, center=True)
            data_cell(ws, row, 6, ev.get("verdict", ""),       bg=bg)
            data_cell(ws, row, 7, ev.get("reasoning") or ev.get("resoning", ""), bg="FFFFFF")

            ws.row_dimensions[row].height = 70

            # Also add to All Evaluations sheet
            run_label = f"{target} / {mode}"
            all_bg    = BG_BUG if bug else BG_CLEAN
            data_cell(ws_all, all_row, 1, run_label,                    bg=BG_ALT, bold=True)
            data_cell(ws_all, all_row, 2, target,                       bg=BG_ALT)
            data_cell(ws_all, all_row, 3, mode,                         bg=BG_ALT)
            data_cell(ws_all, all_row, 4, ev.get("test_id", ""),        bg=BG_ALT)
            data_cell(ws_all, all_row, 5, ev.get("endpoint", ""),       bg=all_bg)
            data_cell(ws_all, all_row, 6, "YES" if bug else "no",       bg=all_bg, bold=bug, center=True)
            data_cell(ws_all, all_row, 7, ev.get("verdict", ""),        bg=all_bg)
            data_cell(ws_all, all_row, 8, ev.get("reasoning") or ev.get("resoning", ""), bg="FFFFFF")
            ws_all.row_dimensions[all_row].height = 70
            all_row += 1

        # Column widths per-sheet
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 13
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 38
        ws.column_dimensions["G"].width = 60

        print(f"  {sheet_name}: {len(evaluations)} evaluations")

    # All Evaluations column widths
    ws_all.column_dimensions["A"].width = 18
    ws_all.column_dimensions["B"].width = 12
    ws_all.column_dimensions["C"].width = 10
    ws_all.column_dimensions["D"].width = 14
    ws_all.column_dimensions["E"].width = 30
    ws_all.column_dimensions["F"].width = 13
    ws_all.column_dimensions["G"].width = 38
    ws_all.column_dimensions["H"].width = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"Total evaluation rows: {all_row - 3}")


if __name__ == "__main__":
    main()