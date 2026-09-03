"""
metrics/export_excel.py
Export thesis comparison metrics to a formatted Excel workbook.
Run from the project root:  python metrics/export_excel.py
"""

import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────

RESULTS_FILE   = Path("results/metrics_summary.json")
OUTPUT_FILE    = Path("results/thesis_metrics.xlsx")

USD_TO_EUR     = 0.8631   # 1 USD = 0.8631 EUR  (2 Sep 2026)
BLENDED_PER_M  = 6.60     # $6.60 / 1M tokens (70% input / 30% output)

# ── Colour palette ───────────────────────────────────────────────────────────

CLR_HEADER_DARK  = "1F3864"   # dark navy
CLR_HEADER_MID   = "2E5DA6"   # medium blue
CLR_AGENT        = "D6E4F7"   # light blue — AI Agent rows
CLR_TRAD         = "EAF4EA"   # light green — traditional framework rows
CLR_ACCENT       = "FFF2CC"   # yellow — highlight cells
CLR_WHITE        = "FFFFFF"
CLR_ALT          = "F5F5F5"   # alternating row grey

FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_TITLE  = Font(name="Arial", bold=True, size=14)
FONT_BODY   = Font(name="Arial", size=10)
FONT_BOLD   = Font(name="Arial", bold=True, size=10)
FONT_SMALL  = Font(name="Arial", size=9, italic=True)

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium() -> Border:
    s = Side(style="medium", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# ── Helpers ──────────────────────────────────────────────────────────────────

def usd(tokens: int) -> float:
    return (tokens / 1_000_000) * BLENDED_PER_M

def eur(tokens: int) -> float:
    return usd(tokens) * USD_TO_EUR

def write_header_row(ws, row: int, values: list, col_start: int = 1,
                     bg: str = CLR_HEADER_DARK):
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        c.font      = FONT_HEADER
        c.fill      = fill(bg)
        c.alignment = CENTER
        c.border    = border_thin()

def write_data_row(ws, row: int, values: list, col_start: int = 1,
                   bg: str = CLR_WHITE, bold: bool = False):
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        c.font      = FONT_BOLD if bold else FONT_BODY
        c.fill      = fill(bg)
        c.alignment = CENTER if isinstance(val, (int, float)) else LEFT
        c.border    = border_thin()

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def section_title(ws, row: int, text: str, cols: int, col_start: int = 1):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_start + cols - 1
    )
    c = ws.cell(row=row, column=col_start, value=text)
    c.font      = Font(name="Arial", bold=True, size=12, color=CLR_HEADER_DARK)
    c.fill      = fill(CLR_ACCENT)
    c.alignment = LEFT
    c.border    = border_medium()

# ── Load data ────────────────────────────────────────────────────────────────

def load() -> list[dict]:
    with open(RESULTS_FILE) as f:
        return json.load(f)["results"]

# ── Sheet 1 — Overview ───────────────────────────────────────────────────────

def sheet_overview(wb, results: list[dict]):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "AI Agentic Software Testing — Thesis Experiment Results"
    t.font      = Font(name="Arial", bold=True, size=16, color=CLR_HEADER_DARK)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d')}  |  "
        f"Model: Claude Sonnet 4-6  |  "
        f"Pricing: $3/$15 per 1M tokens  |  "
        f"Rate: 1 USD = {USD_TO_EUR} EUR"
    )
    s.font      = FONT_SMALL
    s.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # Stats boxes
    agent   = [r for r in results if r["framework"] == "AI Agent"]
    total_t = sum(r["token_usage"] for r in agent)
    total_e = eur(total_t)
    total_b = sum(r["bugs_detected"] for r in agent)

    boxes = [
        ("Total agent runs",        len(agent)),
        ("Total tokens used",       f"{total_t:,}"),
        ("Total cost (EUR)",         f"€{total_e:.4f}"),
        ("Total bugs detected",      total_b),
        ("Frameworks compared",      len(set(r["framework"] for r in results))),
        ("Apps tested",              len(set(r["app"] for r in results))),
    ]
    for i, (label, value) in enumerate(boxes):
        col = 1 + i * 1   # one column per box, we'll spread them
        r4 = ws.cell(row=4, column=col, value=label)
        r4.font = Font(name="Arial", bold=True, size=9, color="666666")
        r4.alignment = CENTER
        r4.fill = fill(CLR_ALT)
        r4.border = border_thin()

        r5 = ws.cell(row=5, column=col, value=value)
        r5.font = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_DARK)
        r5.alignment = CENTER
        r5.fill = fill(CLR_WHITE)
        r5.border = border_thin()

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 28

    # Legend
    row = 7
    ws.merge_cells(f"A{row}:H{row}")
    lg = ws.cell(row=row, column=1, value="Colour Legend")
    lg.font = FONT_BOLD
    lg.fill = fill(CLR_ACCENT)
    lg.alignment = LEFT
    lg.border = border_thin()

    legend = [
        (CLR_AGENT, "AI Agent rows"),
        (CLR_TRAD,  "Traditional framework rows"),
        (CLR_ACCENT,"Highlight / section title"),
    ]
    for j, (color, desc) in enumerate(legend):
        c1 = ws.cell(row=row+1+j, column=1, value="")
        c1.fill   = fill(color)
        c1.border = border_thin()
        c2 = ws.cell(row=row+1+j, column=2, value=desc)
        c2.font   = FONT_BODY
        c2.alignment = LEFT
        c2.border = border_thin()

    set_col_widths(ws, {
        "A": 22, "B": 22, "C": 18, "D": 18,
        "E": 18, "F": 18, "G": 18, "H": 18,
    })

# ── Sheet 2 — Full Comparison ────────────────────────────────────────────────

def sheet_comparison(wb, results: list[dict]):
    ws = wb.create_sheet("Comparison Table")
    ws.sheet_view.showGridLines = False

    headers = [
        "Framework", "App", "Mode",
        "Tests Total", "Tests Passed", "Bugs Detected",
        "False Positives", "Tokens Used"
    ]

    # Freeze pane under header
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    t = ws["A1"]
    t.value     = "Framework Comparison — All Results"
    t.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_DARK)
    t.fill      = fill(CLR_ACCENT)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 22

    write_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 20

    trad_fw = {"Robot Framework", "Pytest + HTTPX", "Schemathesis"}

    for i, r in enumerate(results):
        row   = i + 3
        bg    = CLR_AGENT if r["framework"] == "AI Agent" else CLR_TRAD
        bg    = bg if i % 2 == 0 else (CLR_ALT if r["framework"] != "AI Agent" else "C9DCF0")
        write_data_row(ws, row, [
            r["framework"],
            r["app"],
            r["mode"],
            r.get("tests_executed", 0),
            r.get("tests_passed", 0),
            r["bugs_detected"],
            r["false_positives"],
            r["token_usage"] if r["token_usage"] > 0 else "N/A",
        ], bg=bg)
        ws.row_dimensions[row].height = 16

    set_col_widths(ws, {
        "A": 20, "B": 12, "C": 10,
        "D": 12, "E": 13, "F": 14,
        "G": 14, "H": 14,
    })

# ── Sheet 3 — Bug Detection Summary ─────────────────────────────────────────

def sheet_bugs(wb, results: list[dict]):
    ws = wb.create_sheet("Bug Detection")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "Bug Detection Summary by Application"
    t.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_DARK)
    t.fill      = fill(CLR_ACCENT)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 22

    write_header_row(ws, 2, [
        "App", "Framework", "Mode",
        "Bugs Detected", "False Positives", "FP Rate %"
    ])
    ws.row_dimensions[2].height = 20

    row = 3
    all_apps = sorted(set(r["app"] for r in results))
    for app in all_apps:
        app_results = [r for r in results if r["app"] == app]
        # Section label
        ws.merge_cells(f"A{row}:F{row}")
        sc = ws.cell(row=row, column=1, value=f"  {app.upper()} APP")
        sc.font      = Font(name="Arial", bold=True, size=10, color=CLR_HEADER_DARK)
        sc.fill      = fill(CLR_ACCENT)
        sc.alignment = LEFT
        sc.border    = border_thin()
        ws.row_dimensions[row].height = 18
        row += 1

        for r in app_results:
            bugs = r["bugs_detected"]
            fp   = r["false_positives"]
            total = bugs + fp
            fp_rate = round((fp / total * 100), 1) if total > 0 else 0
            bg = CLR_AGENT if r["framework"] == "AI Agent" else CLR_TRAD
            write_data_row(ws, row, [
                r["app"], r["framework"], r["mode"],
                bugs, fp, f"{fp_rate}%"
            ], bg=bg)
            ws.row_dimensions[row].height = 16
            row += 1

    set_col_widths(ws, {
        "A": 14, "B": 20, "C": 10,
        "D": 15, "E": 15, "F": 12,
    })

# ── Sheet 4 — Agent Efficiency ───────────────────────────────────────────────

def sheet_efficiency(wb, results: list[dict]):
    ws = wb.create_sheet("Agent Efficiency")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = "AI Agent — Efficiency and Cost Analysis"
    t.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_DARK)
    t.fill      = fill(CLR_ACCENT)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 22

    write_header_row(ws, 2, [
        "App", "Mode", "Tests Executed", "Bugs Detected",
        "Tokens Used", "Cost USD", "Cost EUR",
        "Bugs / 1k Tokens", "€ per Bug"
    ])
    ws.row_dimensions[2].height = 20

    agent = [r for r in results if r["framework"] == "AI Agent"]

    for i, r in enumerate(sorted(agent, key=lambda x: (x["app"], x["mode"]))):
        row    = i + 3
        tokens = r["token_usage"]
        bugs   = r["bugs_detected"]
        cost_u = usd(tokens)
        cost_e = eur(tokens)
        eff    = round(bugs / (tokens / 1000), 2) if tokens > 0 else 0
        cpb    = round(cost_e / bugs, 4) if bugs > 0 else 0

        bg = CLR_AGENT if i % 2 == 0 else "C9DCF0"
        write_data_row(ws, row, [
            r["app"], r["mode"],
            r.get("tests_executed", 0), bugs,
            tokens,
            f"${cost_u:.4f}", f"€{cost_e:.4f}",
            eff, f"€{cpb:.4f}"
        ], bg=bg)
        ws.row_dimensions[row].height = 16

    # Totals row
    total_row = len(agent) + 3
    total_t = sum(r["token_usage"] for r in agent)
    total_b = sum(r["bugs_detected"] for r in agent)
    write_data_row(ws, total_row, [
        "TOTAL", "all",
        sum(r.get("tests_executed", 0) for r in agent),
        total_b,
        total_t,
        f"${usd(total_t):.4f}", f"€{eur(total_t):.4f}",
        round(total_b / (total_t / 1000), 2) if total_t > 0 else 0,
        f"€{eur(total_t)/total_b:.4f}" if total_b > 0 else "N/A"
    ], bg=CLR_ACCENT, bold=True)
    ws.row_dimensions[total_row].height = 18

    # Pricing note
    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:I{note_row}")
    n = ws.cell(row=note_row, column=1,
                value=(
                    f"Pricing: Claude Sonnet 4-6 @ $3.00/M input + $15.00/M output tokens  |  "
                    f"Blended estimate: ${BLENDED_PER_M}/M total tokens (70% input / 30% output)  |  "
                    f"Exchange rate: 1 USD = {USD_TO_EUR} EUR (2 Sep 2026, source: Wise)"
                ))
    n.font      = FONT_SMALL
    n.alignment = LEFT

    set_col_widths(ws, {
        "A": 14, "B": 10, "C": 14, "D": 14,
        "E": 14, "F": 12, "G": 12, "H": 16, "I": 12,
    })

# ── Sheet 5 — OS Projects ────────────────────────────────────────────────────

def sheet_os(wb, results: list[dict]):
    ws = wb.create_sheet("OS Projects")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "Open Source Project Testing — Ecological Validity"
    t.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_DARK)
    t.fill      = fill(CLR_ACCENT)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 22

    write_header_row(ws, 2, [
        "Project", "Framework", "Mode",
        "Tests Executed", "Bugs Detected",
        "False Positives", "Tokens"
    ])
    ws.row_dimensions[2].height = 20

    os_results = [r for r in results if "_os" in r["app"]]

    for i, r in enumerate(os_results):
        row = i + 3
        bg  = CLR_AGENT if r["framework"] == "AI Agent" else CLR_TRAD
        write_data_row(ws, row, [
            r["app"], r["framework"], r["mode"],
            r.get("tests_executed", 0),
            r["bugs_detected"],
            r["false_positives"],
            r["token_usage"] if r["token_usage"] > 0 else "N/A",
        ], bg=bg)
        ws.row_dimensions[row].height = 16

    # Note
    note = len(os_results) + 4
    ws.merge_cells(f"A{note}:G{note}")
    n = ws.cell(row=note, column=1,
                value=(
                    "java_os = Swagger Petstore (public API, petstore3.swagger.io)  |  "
                    "python_os = FastAPI Full Stack Template (github.com/fastapi/full-stack-fastapi-template)"
                ))
    n.font      = FONT_SMALL
    n.alignment = LEFT

    set_col_widths(ws, {
        "A": 16, "B": 20, "C": 10,
        "D": 14, "E": 14, "F": 14, "G": 14,
    })

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    results = load()
    print(f"Loaded {len(results)} result records")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    sheet_overview(wb, results)
    sheet_comparison(wb, results)
    sheet_bugs(wb, results)
    sheet_efficiency(wb, results)
    sheet_os(wb, results)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()